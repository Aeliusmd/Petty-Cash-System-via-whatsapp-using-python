"""
Claim Service
OOP class for claim business logic
"""

from typing import Optional, List, Dict, Any
from datetime import date
from fastapi import Request
from app.services.base import BaseService
from app.services.audit_service import AuditService
from app.models.claim import Claim
from app.models.employee import Employee
import uuid
import hashlib
from pathlib import Path
from app.services.textract_service import extract_expense_from_image
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill


class ClaimService(BaseService):
    """
    Service for claim business logic.
    Handles claim CRUD, approvals, rejections, appeals with audit logging.
    """
    
    def __init__(self, request: Optional[Request] = None, audit_service: AuditService = None):
        """Initialize with optional request and audit service"""
        super().__init__(request)
        self.audit_service = audit_service or AuditService(request)
    
    async def get_claim(self, claim_id: int) -> Optional[Claim]:
        """Get claim by ID with full details"""
        return await Claim.find_by_id(claim_id)
    
    async def get_claim_by_number(self, claim_number: str) -> Optional[Claim]:
        """Get claim by claim number"""
        return await Claim.find_by_number(claim_number)
    
    async def get_claims_by_employee(
        self,
        employee_id: int,
        status: str = None,
        limit: int = 10,
        offset: int = 0,
        organization_id: int = None,
        unit_id: int = None,
        start_date: date = None,
        end_date: date = None
    ) -> Dict[str, Any]:
        """Get claims for an employee, filtered by organization. Returns {'claims': [], 'total': int}"""
        claims, total = await Claim.find_by_employee(employee_id, status, limit, offset, organization_id, unit_id, start_date, end_date)
        return {'claims': claims, 'total': total}
    
    async def get_pending_claims_for_manager(self, manager_id: int, organization_id: int = None) -> List[Claim]:
        """Get pending claims for a manager, filtered by organization"""
        return await Claim.find_pending_for_manager(manager_id, organization_id)
    
    async def create_claim(
        self,
        data: Dict[str, Any],
        created_by: int = None,
        organization_id: int = None
    ) -> Claim:
        """Create new claim with audit logging"""
        claim = await Claim.create(data)
        await Claim.initialize_approval_workflow(claim.id)
        
        # Record initial status
        await Claim.record_status_change(
            claim.id, None, 'PENDING', created_by or data.get('employee_id'),
            'Claim submitted'
        )
        
        await self.audit_service.log_claim_action(
            claim_id=claim.id,
            action='CREATE',
            new_values=data,
            performed_by=created_by or data.get('employee_id'),
            organization_id=organization_id
        )
        
        return claim
    
    async def approve_claim(
        self,
        claim_id: int,
        approver_id: int,
        final_amount: float = None,
        organization_id: int = None
    ) -> Optional[Claim]:
        """Approve a claim with audit logging"""
        current = await Claim.find_by_id(claim_id)
        if not current:
            return None
        
        old_values = {
            'status_code': current.status_code,
            'final_amount': current.final_amount
        }
        
        workflow_result = await Claim.approve_step(claim_id, approver_id, final_amount)
        claim = workflow_result["claim"]
        
        # Record status change
        next_assignee_id = workflow_result.get("next_assignee_id")
        if workflow_result.get("is_final"):
            await Claim.record_status_change(
                claim_id, current.status_code, 'APPROVED', approver_id,
                f'Claim approved. Final amount: {claim.final_amount}'
            )
        else:
            await Claim.record_status_change(
                claim_id, current.status_code, current.status_code, approver_id,
                f"Step {workflow_result.get('approved_step')} approved. Routed to next approver."
            )
        
        await self.audit_service.log_claim_action(
            claim_id=claim_id,
            action='APPROVE',
            old_values=old_values,
            new_values={
                'status_code': claim.status_code,
                'final_amount': claim.final_amount,
                'approved_by': approver_id,
                'next_assignee_id': next_assignee_id
            },
            performed_by=approver_id,
            organization_id=organization_id
        )
        
        return claim
    
    async def reject_claim(
        self,
        claim_id: int,
        approver_id: int,
        reason: str,
        organization_id: int = None
    ) -> Optional[Claim]:
        """Reject a claim with audit logging"""
        current = await Claim.find_by_id(claim_id)
        if not current:
            return None
        
        old_values = {
            'status_code': current.status_code,
            'rejection_reason': current.rejection_reason
        }
        
        workflow_result = await Claim.reject_step(claim_id, approver_id, reason)
        claim = workflow_result["claim"]
        
        # Record status change
        await Claim.record_status_change(
            claim_id, current.status_code, 'REJECTED', approver_id,
            f'Rejected: {reason}'
        )
        
        await self.audit_service.log_claim_action(
            claim_id=claim_id,
            action='REJECT',
            old_values=old_values,
            new_values={
                'status_code': 'REJECTED',
                'rejection_reason': reason,
                'approved_by': approver_id
            },
            performed_by=approver_id,
            organization_id=organization_id
        )
        
        return claim
    
    async def appeal_claim(
        self,
        claim_id: int,
        employee_id: int,
        notes: str = None,
        organization_id: int = None
    ) -> Optional[Claim]:
        """Appeal a rejected claim with audit logging"""
        current = await Claim.find_by_id(claim_id)
        if not current:
            return None
        
        if current.status_code != 'REJECTED':
            raise ValueError("Only rejected claims can be appealed")
        
        old_values = {
            'status_code': current.status_code,
            'appeal_count': current.appeal_count
        }
        
        claim = await Claim.appeal(claim_id, employee_id, notes)
        
        await self.audit_service.log_claim_action(
            claim_id=claim_id,
            action='APPEAL',
            old_values=old_values,
            new_values={
                'status_code': 'APPEALED',
                'appeal_count': claim.appeal_count if claim else old_values['appeal_count'] + 1
            },
            performed_by=employee_id,
            organization_id=organization_id,
            metadata={'notes': notes}
        )
        
        return claim
    
    async def get_claim_stats(self, employee_id: int) -> Dict[str, Any]:
        """Get claim statistics for an employee"""
        return await Claim.get_stats(employee_id)
    
    async def get_claim_history(self, claim_id: int) -> List[Dict[str, Any]]:
        """Get status history for a claim"""
        return await Claim.get_history(claim_id)
    
    async def add_receipt(
        self,
        claim_id: int,
        file_path: str,
        file_name: str,
        file_type: str = None,
        file_size: int = None,
        ocr_amount: float = None,
        ocr_raw_text: str = None,
        message_id: str = None,
        vendor: str = None,
        uploaded_by: int = None,
        organization_id: int = None
    ) -> Dict[str, Any]:
        """Add receipt to claim with audit logging"""
        receipt = await Claim.add_receipt(
            claim_id, file_path, file_name, file_type,
            file_size, ocr_amount, ocr_raw_text, message_id, vendor
        )
        
        if receipt and uploaded_by:
            await self.audit_service.log_receipt_action(
                receipt_id=receipt.get('id'),
                claim_id=claim_id,
                action='UPLOAD',
                metadata={
                    'file_name': file_name,
                    'file_type': file_type,
                    'ocr_amount': ocr_amount,
                    'vendor': vendor
                },
                performed_by=uploaded_by,
                organization_id=organization_id
            )
        
        return receipt
    
    async def get_receipts(self, claim_id: int) -> List[Dict[str, Any]]:
        """Get all receipts for a claim"""
        return await Claim.get_receipts(claim_id)
    
    async def process_and_save_receipts(
        self,
        claim_id: int,
        receipt_files: List[Dict[str, Any]],
        employee_id: int,
        organization_id: int = None
    ) -> Dict[str, Any]:
        """
        Process multiple receipts: extract OCR, save to DB, calculate total.
        Shared method for both WhatsApp and Web uploads.
        
        Args:
            claim_id: ID of the claim to attach receipts to
            receipt_files: List of dicts with keys:
                - 'bytes': file content as bytes
                - 'filename': original filename
                - 'content_type': MIME type
            employee_id: ID of employee uploading
            organization_id: Organization context
            
        Returns:
            {
                'receipts_saved': int,
                'total_extracted': float,
                'receipts': [list of receipt records]
            }
        """
        from app.services import textract_service
        from pathlib import Path
        import uuid
        
        total_extracted = 0.0
        receipts_saved = []
        warnings = []
        saved_count = 0
        
        # 1. Get receipts directory (handle both local and Docker environments)
        possible_paths = [
            Path("/app/backend/receipts"),  # Docker container
            Path(__file__).resolve().parent.parent / "receipts",  # Local relative to app folder
            Path("receipts").resolve()  # Current working directory
        ]
        
        receipts_dir = None
        for path in possible_paths:
            if path.exists():
                receipts_dir = path
                break
        
        if not receipts_dir:
            # Fallback: create relative to this file's location
            receipts_dir = Path(__file__).resolve().parent.parent / "receipts"
            receipts_dir.mkdir(parents=True, exist_ok=True)
        
        for receipt_data in receipt_files:
            content = receipt_data['bytes']
            filename = receipt_data.get('filename', 'receipt')
            content = receipt_data.get('bytes')
            filename = receipt_data.get('filename')
            content_type = receipt_data.get('content_type')
            
            if not content:
                continue
                
            # 1. Save to disk
            ext = filename.split('.')[-1] if '.' in filename else 'jpg'
            saved_filename = f"{uuid.uuid4()}.{ext}"
            file_path = receipts_dir / saved_filename
            
            with open(file_path, "wb") as f:
                f.write(content)
            
            # 2. Perform OCR / Extraction
            ocr_amount = 0.0
            ocr_text = ""
            ocr_vendor = ""
            ocr_date = ""
            ocr_invoice_id = ""
            
            try:
                # Use textract service
                result = await extract_expense_from_image(content)
                if result and result.get('success'):
                    extracted = result.get('expense', {})
                    ocr_amount = float(extracted.get('total', 0) or 0)
                    ocr_vendor = extracted.get('vendorName', '')
                    ocr_date = extracted.get('date', '')
                    ocr_invoice_id = extracted.get('invoiceId', '')
                    
                    # Use the Textract-generated text directly — it already contains
                    # ALL summary fields as "Label: Value" lines (vendor, date, total,
                    # invoice ID, subtotal, tax, etc.) ready for the frontend to parse.
                    ocr_text = result.get('text', '')
                    
                    if ocr_amount > 0:
                        total_extracted += ocr_amount
            except Exception as e:
                print(f"OCR failed for {filename}: {e}")
            
            # 3. Calculate OCR Content Fingerprint for Duplicate Detection
            # Uses extracted data, not raw bytes (invariant to re-compression)
            file_hash = None
            fingerprint_parts = [
                (ocr_vendor or '').strip().lower(),
                str(ocr_amount or '').strip(),
                (ocr_date or '').strip(),
                str(ocr_invoice_id or '').strip()
            ]
            fingerprint_str = '|'.join(fingerprint_parts)
            if any(fingerprint_parts):
                file_hash = hashlib.sha256(fingerprint_str.encode()).hexdigest()
                print(f"🔑 Content fingerprint: {fingerprint_str} -> {file_hash[:16]}...")
            
            # Check for duplicates
            if file_hash:
                duplicate = await Claim.check_duplicate_receipt(file_hash, current_claim_id=claim_id)
                if duplicate:
                    print(f"⚠️ Duplicate receipt detected! Used in claim #{duplicate.get('claim_number')}")
                    warnings.append({
                        'type': 'duplicate',
                        'message': f"Potential duplicate: Used in Claim #{duplicate.get('claim_number')} by {duplicate.get('employee_name')} on {duplicate.get('created_at').strftime('%Y-%m-%d') if duplicate.get('created_at') else 'unknown date'}",
                        'original_claim': duplicate
                    })
            
            # 4. Save to database
            receipt_record = await Claim.add_receipt(
                claim_id=claim_id,
                file_path=saved_filename,
                file_name=filename,
                file_type=content_type,
                file_size=len(content),
                ocr_amount=ocr_amount,
                ocr_raw_text=ocr_text,
                vendor=ocr_vendor,
                file_hash=file_hash
            )
            
            if receipt_record:
                saved_count += 1
        
        # Recalculate system amount
        if saved_count > 0:
            await Claim.calculate_system_amount(claim_id)
                
        return {
            "total_extracted": total_extracted,
            "receipt_count": saved_count,
            "warnings": warnings
        }
    
    async def export_claims(
        self,
        format: str = 'csv',
        employee_id: int = None, 
        unit_id: int = None, 
        status: str = None, 
        start_date: date = None, 
        end_date: date = None,
        organization_id: int = None
    ) -> Any:
        """
        Export claims to CSV or Excel
        Returns: BytesIO object containing the file
        """
        # Fetch data
        claims = await Claim.find_for_export(
            employee_id=employee_id,
            unit_id=unit_id,
            status=status,
            start_date=start_date,
            end_date=end_date,
            organization_id=organization_id
        )
        
        # Define output headers
        headers = [
            'Claim Number', 'Date', 'Employee Name', 'Employee Code', 'Department',
            'Category', 'Description', 'User Amount', 'Total from Receipts', 'Difference', 'Approved Amount', 'Status', 'Approved By'
        ]
        
        output = io.BytesIO()
        
        if format.lower() == 'csv':
            # Create text buffer for CSV
            text_buffer = io.StringIO()
            writer = csv.writer(text_buffer)
            writer.writerow(headers)
            
            for claim in claims:
                user_amt = float(claim.user_amount) if claim.user_amount is not None else 0.0
                rec_total = getattr(claim, 'receipt_total', 0.0) or 0.0
                diff = abs(user_amt - rec_total)
                # Check status
                is_approved = claim.status_code == 'APPROVED'
                final_amt = float(claim.final_amount) if claim.final_amount is not None else 0.0
                
                writer.writerow([
                    claim.claim_number,
                    claim.claim_date.strftime('%Y-%m-%d') if claim.claim_date else '',
                    claim.employee_name,
                    claim.employee_code,
                    claim.unit_name or '',
                    claim.category_name,
                    claim.description or '',
                    f"{user_amt:.2f}",
                    f"{rec_total:.2f}",
                    f"{diff:.2f}",
                    f"{final_amt:.2f}" if is_approved else '',
                    claim.status_name,
                    claim.approver_name if is_approved and getattr(claim, 'approver_name', None) else ''
                ])
                
            # Convert to bytes
            output.write(text_buffer.getvalue().encode('utf-8'))
            output.seek(0)
            return output
            
        elif format.lower() == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Claims Export"
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Write data
            for row_num, claim in enumerate(claims, 2):
                user_amt = float(claim.user_amount) if claim.user_amount is not None else 0.0
                rec_total = getattr(claim, 'receipt_total', 0.0) or 0.0
                diff = abs(user_amt - rec_total)
                # Check status
                is_approved = claim.status_code == 'APPROVED'
                final_amt = float(claim.final_amount) if claim.final_amount is not None else 0.0

                ws.cell(row=row_num, column=1, value=claim.claim_number)
                ws.cell(row=row_num, column=2, value=claim.claim_date.strftime('%Y-%m-%d') if claim.claim_date else '')
                ws.cell(row=row_num, column=3, value=claim.employee_name)
                ws.cell(row=row_num, column=4, value=claim.employee_code)
                ws.cell(row=row_num, column=5, value=claim.unit_name or '')
                ws.cell(row=row_num, column=6, value=claim.category_name)
                ws.cell(row=row_num, column=7, value=claim.description or '')
                ws.cell(row=row_num, column=8, value=user_amt)
                ws.cell(row=row_num, column=9, value=rec_total)
                ws.cell(row=row_num, column=10, value=diff)
                ws.cell(row=row_num, column=11, value=final_amt if is_approved else '')
                ws.cell(row=row_num, column=12, value=claim.status_name)
                ws.cell(row=row_num, column=13, value=claim.approver_name if is_approved and getattr(claim, 'approver_name', None) else '')
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
                
            wb.save(output)
            output.seek(0)
            return output
            
        else:
            raise ValueError("Unsupported format. Use 'csv' or 'xlsx'")

    async def delete_claim(
        self,
        claim_id: int,
        deleted_by: int,
        organization_id: int = None
    ) -> bool:
        """Delete a claim with audit logging"""
        current = await Claim.find_by_id(claim_id)
        if not current:
            return False
        
        success = await Claim.delete(claim_id)
        
        if success:
            await self.audit_service.log_claim_action(
                claim_id=claim_id,
                action='DELETE',
                old_values=current.to_dict(),
                performed_by=deleted_by,
                organization_id=organization_id
            )
        
        return success
